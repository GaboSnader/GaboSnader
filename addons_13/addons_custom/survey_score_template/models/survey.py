# -*- coding:utf-8 -*-
from odoo import models, fields, api
from datetime import date, datetime
import base64
from base64 import b64encode
from werkzeug import urls

from tabulate import tabulate
import openpyxl
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from pandas import DataFrame, crosstab
from openpyxl.styles import Font, PatternFill, Color, Alignment, Border, Side
from openpyxl.styles.borders import BORDER_THICK
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.worksheet.hyperlink import Hyperlink

titleFill = PatternFill(start_color='000CA4', end_color='000CA4', fill_type='solid')
blueFill = PatternFill(start_color='8AEAFB', end_color='8AEAFB', fill_type='solid')
blue2Fill = PatternFill(start_color='3DE2FF', end_color='3DE2FF', fill_type='solid')
blue3Fill = PatternFill(start_color='8087FF', end_color='8087FF', fill_type='solid')
lowFill = PatternFill(start_color='DAF7A6', end_color='DAF7A6', fill_type='solid')
mediumFill = PatternFill(start_color='FFC300', end_color='FFC300', fill_type='solid')
highFill = PatternFill(start_color='FF5733', end_color='FF5733', fill_type='solid')
veryFill = PatternFill(start_color='C70039', end_color='C70039', fill_type='solid')
blueFont = Font(name='Arial', size=15, color='000CA4')
whiteFont = Font(color='FFFFFF')

class SurveySurvey(models.Model):
    _inherit = 'survey.survey'

    qualification_id = fields.Many2one('survey_score_template.qualification',string=u'Calificación')
    sequence = fields.Integer('Secuencia')
    date_start = fields.Date(string='Fecha de inicio')
    survey_url = fields.Char("Public link", compute="_compute_survey_url")

    def _compute_survey_url(self):
        """ Computes a public URL for the survey """
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        for survey in self:
            survey.survey_url = urls.url_join(base_url, "survey/nom-35/")
            survey.public_url = urls.url_join(base_url, "survey/start/%s" % (survey.access_token))


    def open_print_report_xls(self):
        headers = self._get_header()
        details = self.generate_data(self.user_input_ids)
        filename = 'Escuesta.xlsx' 
        output = self._create_xlsx(
            Workbook, NamedTemporaryFile, headers=headers, body=details)
        xlsx = {
            'name': filename,
            # 'datas_fname': filename,
            'type': 'binary',
            'res_model': self._name,
            'datas': base64.b64encode(output),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        inserted_id = self.env['ir.attachment'].create(xlsx)
        url = '/web/content/%s?download=1' % (inserted_id.id)
        return {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': url
        }

    def generate_data(self, users):
        res = []
        for user in users:
            indiv = 0.0
            result = ''
            for line in user.user_input_line_ids:
                indiv += line.answer_score
            low = self.qualification_id.low
            medium = self.qualification_id.medium
            high = self.qualification_id.high
            result, fill = self._get_result(indiv, low, medium, high)
            res.append({
                'user': user.partner_id.name,
                'result': result,
                'fill': fill,
                'answer_score': indiv,
            })
        return res

    def _get_usr(self, user):
        indiv = 0.0
        for line in user.user_input_line_ids:
            indiv += line.answer_score
        return indiv

    def _get_res_emp(self, employ, integer):
        if integer > 0:
            usr = self.env['survey.user_input']
            for emp in employ:
                usr += self.env['survey.user_input'].search([('partner_id.name','=',emp.name)])
            result = 0.0
            if usr:
                result += self._get_usr(usr)
                return int(result/len(usr))
            else:
                return 0
        else:
            return 0

    def _get_res_usr(self, user, integer):
        if integer > 0:
            usr = self.env['survey.user_input']
            for emp in user:
                usr += self.env['survey.user_input'].search([('partner_id.name','=',emp.user_id.name)])
            result = 0.0
            if usr:
                result += self._get_usr(usr)
                return int(result/len(usr))
            else:
                return 0
        else:
            return 0

    def _get_result(self, indiv, low, medium, high):
        result = ''
        fill = False
        if not indiv:
            result = 'NULO'
            fill = blueFill
        elif indiv <= low:
            result = 'BAJO'
            fill = lowFill
        elif indiv <= medium:
            result = 'MEDIO'
            fill = mediumFill
        elif indiv <= high:
            result = 'ALTO'
            fill = highFill
        else:
            result = 'MUY ALTO'
            fill = veryFill
        return result, fill


    def _create_xlsx(self, Workbook, NamedTemporaryFile,headers=None, body=None):
        wb = Workbook()
        ws = wb.active
        self._create_header(ws, headers)
        self._create_body(ws, body)
        self._categ_row(ws, self.user_input_ids)
        self._dom_row(ws, self.user_input_ids)
        self._res_row(ws, self.user_input_ids)
        output = None
        with NamedTemporaryFile() as tmp:
            for column_cells in ws.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
            wb.save(tmp.name)
            output = tmp.read()
        return output

    def _create_header(self, ws, headers):
        ws.cell(row=1,
                column=1,
                value='Evaluado - Resultado Cuestionario').font = blueFont
        ws.cell(row=1,
                column=6+2,
                value='Evaluado - Resultado Categoria').font = blueFont
        ws.cell(row=1,
                column=12+4,
                value='Evaluado - Estructura Dominio/Dimension').font = blueFont
        ws.cell(row=1,
                column=19+6,
                value='Evaluado - Resultado Dominio').font = blueFont
        index = 1
        for header in headers:
            if header:
                cell = ws.cell(row=2, column=index, value=header)
                cell.fill = titleFill
                cell.font = whiteFont
            else:
                ws.cell(row=2, column=index, value=header)
            index += 1
        return ws
    
    def _create_body(self, ws, regs):
        index = 3
        fill = blueFill
        for reg in regs:
            self._create_row(ws, index, reg, fill)
            index += 1
            if fill == blueFill:
                fill = blue2Fill
            else:
                fill = blueFill
        ws.cell(row=index,
                column=2,
                value='Total').fill = blue3Fill
        ws.cell(row=index,
                column=3,
                value=self.answer_done_count).fill = blue3Fill
        return ws

    def _create_row(self, ws, index, reg, fill):
        department = ''
        workplace = ''
        employee_ids = self.env['hr.employee'].search([('name','=',reg['user'])])
        for emp in employee_ids:
            department = str(emp.department_id.name)
            workplace = str(emp.workplace_id.name)
            ws.cell(row=index,
                column=1,
                value=workplace).fill = fill
            ws.cell(row=index,
                column=2,
                value=department).fill = fill
        ws.cell(row=index,
            column=3,
            value=reg['user']).fill = fill
        ws.cell(row=index,
            column=4,
            value=reg['result']).fill = reg['fill']
        ws.cell(row=index,
            column=5,
            value=reg['answer_score']).fill = fill
        return ws

    def _get_employee(self, user):
        employee_ids = self.env['hr.employee'].search([('name','=',user.partner_id.name)])
        for emp in employee_ids:
            return emp

    def _categ_row(self, ws, users):
        index = 3
        index2 = 3
        fill = blueFill
        fill2 = blueFill
        for user in users:
            employee = self._get_employee(user)
            if employee:
                # ws.merge_cells(start_row=index, start_column=8,
                                # end_row=index+4, end_column=8)
                for row in range(index, index+5):
                    ws.cell(row=row,
                        column=8,
                        value=employee.workplace_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=9,
                                # end_row=index+4, end_column=9)
                for row in range(index, index+5):
                    ws.cell(row=row,
                        column=9,
                        value=employee.department_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=8+2,
                                # end_row=index+4, end_column=8+2)
                for row in range(index, index+5):
                    ws.cell(row=row,
                        column=8+2,
                        value=user.partner_id.name).fill = fill
                categ = self._get_categ_data()
                result = False
                for c in categ:
                    val = 0.0
                    ws.cell(row=index2, column=9+2, value=c).fill = fill2
                    low = 0.0
                    medium = 0.0
                    high = 0.0
                    for line in user.user_input_line_ids:
                        low = line.value_suggested_row.category_id.low
                        medium = line.value_suggested_row.category_id.medium
                        high = line.value_suggested_row.category_id.high
                        if line.value_suggested_row.category_id.name == c:
                            val += line.answer_score
                    result, fill = self._get_result(val, low, medium, high)
                    ws.cell(row=index2, column=10+2, value=str(result)).fill = fill
                    ws.cell(row=index2, column=11+2, value=val).fill = fill2
                    index2 += 1
                    if fill2 == blueFill:
                        fill2 = blue2Fill
                    else:
                        fill2 = blueFill
                index += 5
                if fill == blueFill:
                    fill = blue2Fill
                else:
                    fill = blueFill
        return ws

    def _get_categ_data(self):
        return [
            'Liderazgo y relaciones en el trabajo',
            'Factores propios de la actividad',
            'Organización del tiempo de trabajo',
            'Entorno organizacional',
            'Ambiente de trabajo'
        ]

    def _dom_row(self, ws, users):
        index = 3
        index2 = 3
        fill = blueFill
        fill2 = blueFill
        fill3 = blueFill
        fill4 = blue2Fill
        for user in users:
            employee = self._get_employee(user)
            if employee:
                # ws.merge_cells(start_row=index, start_column=16,
                                # end_row=index+24, end_column=16)
                for row in range(index, index+25):
                    ws.cell(row=row,
                        column=16,
                        value=employee.workplace_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=17,
                                # end_row=index+24, end_column=17)
                for row in range(index, index+25):
                    ws.cell(row=row,
                        column=17,
                        value=employee.department_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=12+6,
                                # end_row=index+24, end_column=12+6)
                for row in range(index, index+25):
                    ws.cell(row=row,
                        column=12+6,
                        value=user.partner_id.name).fill = fill3
                # ws.merge_cells(start_row=index, start_column=13+6,
                                # end_row=index+4, end_column=13+6)
                for row in range(index, index+5):
                    ws.cell(row=row,
                        column=13+6,
                        value='Liderazgo y relaciones en el trabajo').fill = fill3
                # ws.merge_cells(start_row=index+5, start_column=13+6,
                                # end_row=index+14, end_column=13+6)
                for row in range(index+5, index+15):
                    ws.cell(row=row,
                        column=13+6,
                        value='Factores propios de la actividad').fill = fill4
                # ws.merge_cells(start_row=index+15, start_column=13+6,
                                # end_row=index+17, end_column=13+6)
                for row in range(index+15, index+18):
                    ws.cell(row=row,
                        column=13+6,
                        value='Organización del tiempo de trabajo').fill = fill3
                # ws.merge_cells(start_row=index+18, start_column=13+6,
                                # end_row=index+21, end_column=13+6)
                for row in range(index+18, index+22):
                    ws.cell(row=row,
                        column=13+6,
                        value='Entorno organizacional').fill = fill4
                # ws.merge_cells(start_row=index+22, start_column=13+6,
                                # end_row=index+24, end_column=13+6)
                for row in range(index+22, index+25):
                    ws.cell(row=row,
                        column=13+6,
                        value='Ambiente de trabajo').fill = fill3
                # ws.merge_cells(start_row=index, start_column=14+6,
                                # end_row=index+1, end_column=14+6)
                for row in range(index, index+2):
                    ws.cell(row=row,
                        column=14+6,
                        value='Liderazgo').fill = fill3
                # ws.merge_cells(start_row=index+2, start_column=14+6,
                                # end_row=index+3, end_column=14+6)
                for row in range(index+2, index+4):
                    ws.cell(row=row,
                        column=14+6,
                        value='Relaciones de trabajo').fill = fill4
                ws.cell(row=index+4,
                        column=14+6,
                        value='Violencia').fill = fill3
                # ws.merge_cells(start_row=index+5, start_column=14+6,
                                # end_row=index+10, end_column=14+6)
                for row in range(index+5, index+11):
                    ws.cell(row=row,
                        column=14+6,
                        value='Carga de trabajo').fill = fill4
                # ws.merge_cells(start_row=index+11, start_column=14+6,
                                # end_row=index+14, end_column=14+6)
                for row in range(index+11, index+15):
                    ws.cell(row=row,
                        column=14+6,
                        value='Falta de control sobre el trabajo').fill = fill3
                # ws.merge_cells(start_row=index+15, start_column=14+6,
                                # end_row=index+16, end_column=14+6)
                for row in range(index+15, index+17):
                    ws.cell(row=row,
                        column=14+6,
                        value='Interferencia en la relación trabajo-familia').fill = fill4
                ws.cell(row=index+17,
                        column=14+6,
                        value='Jornada de trabajo').fill = fill3
                # ws.merge_cells(start_row=index+18, start_column=14+6,
                                # end_row=index+19, end_column=14+6)
                for row in range(index+18, index+20):
                    ws.cell(row=row,
                        column=14+6,
                        value='Insuficiente sentido de pertenencia e inestabilidad').fill = fill4
                # ws.merge_cells(start_row=index+20, start_column=14+6,
                                # end_row=index+21, end_column=14+6)
                for row in range(index+20, index+22):
                    ws.cell(row=row,
                        column=14+6,
                        value='Reconocimiento del desempeño').fill = fill3
                # ws.merge_cells(start_row=index+22, start_column=14+6,
                                # end_row=index+24, end_column=14+6)
                for row in range(index+22, index+25):
                    ws.cell(row=row,
                        column=14+6,
                        value='Condiciones en el ambiente de trabajo').fill = fill4
                dim = self._get_dim()
                for d in dim:
                    val = 0.0
                    ws.cell(row=index2, column=15+6, value=d).fill = fill2
                    for line in user.user_input_line_ids:
                        if line.value_suggested_row.dimension_id.name == d:
                            val += line.answer_score
                    ws.cell(row=index2, column=16+6, value=val). fill = fill2
                    index2 += 1
                    if fill2 == blueFill:
                        fill2 = blue2Fill
                    else:
                        fill2 = blueFill
                if fill == blueFill:
                    fill = blue2Fill
                else:
                    fill = blueFill
                index += 25
                if fill3 == blueFill:
                    fill3 = blue2Fill
                    fill4 = blueFill
                else:
                    fill3 = blueFill
                    fill4 = blue2Fill
        return ws

    def _get_dim(self):
        return [
            'Características del liderazgo',
            'Escaza claridad de funciones',
            'Deficiente relación con los colaboradores que supervisa',
            'Relaciones sociales en el trabajo',
            'Violencia laboral',
            'Carga mental',
            'Cargas contradictorias o incosistentes',
            'Cargas cuantitativas',
            'Cargas de alta responsabilidad',
            'Cargas psicológicas emocionales',
            'Ritmo de trabajo acelerado',
            'Falta de control y autonomía sobre el trabajo',
            'Insuficiente participación y manejo del cambio',
            'Limitada o inexistente capacitación',
            'Limitada o nula posibilidad de desarrollo',
            'Influencia de las responsabilidades familiares',
            'Influencia del trabajo fuera del centro laboral',
            'Jornadas de trabajo extensas',
            'Inestabilidad laboral',
            'Limitado sentido de pertenencia',
            'Escasa o nula retroalimentación del desempeño',
            'Escaso o nulo reconocimiento y compensación',
            'Condiciones deficientes e insalubres',
            'Condiciones peligrosas e inseguras',
            'Trabajos peligrosos',
        ]

    def _res_row(self, ws, users):
        index = 3
        index2 = 3
        fill = blueFill
        fill2 = blueFill
        fill3 = blueFill
        fill4 = blue2Fill
        for user in users:
            employee = self._get_employee(user)
            if employee:
                # ws.merge_cells(start_row=index, start_column=25,
                                # end_row=index+9, end_column=25)
                for row in range(index, index+10):
                    ws.cell(row=row,
                        column=25,
                        value=employee.workplace_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=26,
                                # end_row=index+9, end_column=26)
                for row in range(index, index+10):
                    ws.cell(row=row,
                        column=26,
                        value=employee.department_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=19+8,
                                # end_row=index+9, end_column=19+8)
                for row in range(index, index+10):
                    ws.cell(row=row,
                        column=19+8,
                        value=user.partner_id.name).fill = fill
                # ws.merge_cells(start_row=index, start_column=20+8,
                                # end_row=index+2, end_column=20+8)
                for row in range(index, index+3):
                    ws.cell(row=row,
                        column=20+8,
                        value='Liderazgo y relaciones en el trabajo').fill = fill3
                # ws.merge_cells(start_row=index+3, start_column=20+8,
                                # end_row=index+4, end_column=20+8)
                for row in range(index+3, index+5):
                    ws.cell(row=row,  
                        column=20+8,
                        value='Factores propios de la actividad').fill = fill4
                # ws.merge_cells(start_row=index+5, start_column=20+8,
                                # end_row=index+6, end_column=20+8)
                for row in range(index+5, index+7):
                    ws.cell(row=row,  
                        column=20+8,
                        value='Organización del tiempo de trabajo').fill = fill3
                # ws.merge_cells(start_row=index+7, start_column=20+8,
                                # end_row=index+8, end_column=20+8)
                for row in range(index+7, index+10):
                    ws.cell(row=row,  
                        column=20+8,
                        value='Entorno organizacional').fill = fill4
                ws.cell(row=index+9,
                        column=20+8,
                        value='Ambiente de trabajo').fill = fill3
                dom = self._get_dom()
                for d in dom:
                    val = 0.0
                    ws.cell(row=index2, column=21+8, value=d).fill = fill2
                    low = 0.0
                    medium = 0.0
                    high = 0.0
                    for line in user.user_input_line_ids:
                        low = line.value_suggested_row.domain_id.low
                        medium = line.value_suggested_row.domain_id.medium
                        high = line.value_suggested_row.domain_id.high
                        if line.value_suggested_row.domain_id.name == d:
                            val += line.answer_score
                    result, fill = self._get_result(val, low, medium, high)
                    ws.cell(row=index2, column=22+8, value=result).fill = fill
                    ws.cell(row=index2, column=23+8, value=val).fill = fill2
                    index2 += 1
                    if fill2 == blueFill:
                        fill2 = blue2Fill
                    else:
                        fill2 = blueFill
                if fill == blueFill:
                    fill = blue2Fill
                else:
                    fill = blueFill
                if fill3 == blueFill:
                    fill3 = blue2Fill
                    fill4 = blueFill
                else:
                    fill3 = blueFill
                    fill4 = blue2Fill
                index += 10
        return ws

    def _get_dom(self):
        return [
            'Violencia',
            'Liderazgo',
            'Relaciones de trabajo',
            'Falta de control sobre el trabajo',
            'Carga de trabajo',
            'Interferencia en la relación trabajo-familia',
            'Jornada de trabajo',
            'Insuficiente sentido de pertenencia e inestabilidad',
            'Reconocimiento del desempeño',
            'Condiciones en el ambiente de trabajo',
        ]


    def _get_header(self):
        return [
            'CENTRO DE TRABAJO',
            'DEPARTAMENTO',
            'NOMBRE EVALUADO',
            'RESULTADO CUESTIONARIO',
            'RESULTADO INVIVIDUAL',
            '',
            '',
            'CENTRO DE TRABAJO',
            'DEPARTAMENTO',
            'NOMBRE EVALUADO',
            'CATEGORIA',
            'RESULTADO CATEGORIA',
            'RESULTADO',
            '',
            '',
            'CENTRO DE TRABAJO',
            'DEPARTAMENTO',
            'NOMBRE EVALUADO',
            'CATEGORIA',
            'DOMINIO',
            'DIMENSIÓN',
            'RESPUESTAS',
            '',
            '',
            'CENTRO DE TRABAJO',
            'DEPARTAMENTO',
            'NOMBRE EVALUADO',
            'CATEGORIA',
            'DOMINIO',
            'RESULTADO DOMINIO',
            'RESULTADO',
        ]

    def as_text(self,value):
        if value is None:
            return ""
        return str(value)

    def open_print_report_department_xls(self):
        filename = 'Escuesta Departamento.xlsx' 
        output = self._create_xlsx_department(
            Workbook, NamedTemporaryFile)
        xlsx = {
            'name': filename,
            # 'datas_fname': filename,
            'type': 'binary',
            'res_model': self._name,
            'datas': base64.b64encode(output),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        inserted_id = self.env['ir.attachment'].create(xlsx)
        url = '/web/content/%s?download=1' % (inserted_id.id)
        return {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': url
        }

    def _get_header_department(self):
        return [
            'CENTRO DE TRABAJO',
            'DEPARTAMENTO',
            'RESULTADO',
            'EMPLEADOS',
            '% EMPLEADOS',
        ]

    def _create_xlsx_department(self, Workbook, NamedTemporaryFile):
        wb = Workbook()
        ws = wb.active
        self._create_body_department(ws, self.user_input_ids)
        output = None
        with NamedTemporaryFile() as tmp:
            for column_cells in ws.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
            wb.save(tmp.name)
            output = tmp.read()
        return output

    def _create_body_department(self, ws, users):
        column = 1
        ws.cell(row=1,
                column=1,
                value='Departamento - Distribución Resultado Departamento').font = blueFont
        header = self._get_header_department()
        for c in header:
            cell = ws.cell(row=2, column=column, value=c)
            cell.fill = titleFill
            cell.font = whiteFont
            column += 1
        index = 3
        fill = blueFill
        fill2 = blue2Fill
        workplace_ids = self.env['survey.workplace']
        employee_ids = self.env['hr.employee']
        for user in users:
            employee_ids += self.env['hr.employee'].search([('name','=',user.partner_id.name)])
            for emp in employee_ids:
                if not emp.workplace_id in workplace_ids:
                    workplace_ids += emp.workplace_id
        int_emp = 0
        for workplace in workplace_ids:
            department_ids = self.env['hr.department']
            # emp_ids = employee_ids.search([('workplace_id','=',workplace.id)])
            for emp in employee_ids:
                if not emp.department_id in department_ids and emp.workplace_id.id == workplace.id:
                    department_ids += emp.department_id
            if department_ids:
                # ws.merge_cells(start_row=index, start_column=1,
                #                 end_row=index+(len(department_ids)*5)-1, end_column=1)
                for row in range(index, index+(len(department_ids)*5)):
                    ws.cell(row=row, column=1, value=workplace.name).fill = fill
                for department in department_ids:
                    # ws.merge_cells(start_row=index, start_column=2,
                    #                 end_row=index+4, end_column=2)
                    for row in range(index, index+5):
                        ws.cell(row=row,
                                column=2,
                                value=department.name).fill = fill
                    n_result = 0
                    b_result = 0
                    m_result = 0
                    a_result = 0
                    ma_result = 0
                    no_employee = 0
                    for user in users:
                        for emp in employee_ids:
                            if emp.name == user.partner_id.name and emp.department_id.id == department.id and workplace.id == emp.workplace_id.id:
                                indiv = 0.0
                                for line in user.user_input_line_ids:
                                    indiv += line.answer_score
                                low = self.qualification_id.low
                                medium = self.qualification_id.medium
                                high = self.qualification_id.high
                                result, fill3 = self._get_result(indiv, low, medium, high)
                                if result == 'NULO':
                                    n_result += 1
                                elif result == 'BAJO':
                                    b_result += 1
                                elif result == 'MEDIO':
                                    m_result += 1
                                elif result == 'ALTO':
                                    a_result += 1
                                else:
                                    ma_result += 1
                                no_employee += 1
                                ws.cell(row=index, column=3, value='NULO').fill = blueFill
                                ws.cell(row=index+1, column=3, value='BAJO').fill = lowFill
                                ws.cell(row=index+2, column=3, value='MEDIO').fill = mediumFill
                                ws.cell(row=index+3, column=3, value='ALTO').fill = highFill
                                ws.cell(row=index+4, column=3, value='MUY ALTO').fill = veryFill
                                ws.cell(row=index,
                                        column=4,
                                        value=n_result).fill = fill
                                ws.cell(row=index+1,
                                        column=4,
                                        value=b_result).fill = fill2
                                ws.cell(row=index+2,
                                        column=4,
                                        value=m_result).fill = fill
                                ws.cell(row=index+3,
                                        column=4,
                                        value=a_result).fill = fill2
                                ws.cell(row=index+4,
                                        column=4,
                                        value=ma_result).fill = fill
                                try:
                                    ws.cell(row=index,
                                            column=5,
                                            value=(n_result*100/no_employee)).fill = fill
                                except:
                                    ws.cell(row=index,
                                            column=5,
                                            value=0).fill = fill
                                try:
                                    ws.cell(row=index+1,
                                            column=5,
                                            value=(b_result*100/no_employee)).fill = fill2
                                except:
                                    ws.cell(row=index+1,
                                            column=5,
                                            value=0).fill = fill2
                                try:
                                    ws.cell(row=index+2,
                                            column=5,
                                            value=(m_result*100/no_employee)).fill = fill
                                except:
                                    ws.cell(row=index+2,
                                            column=5,
                                            value=0).fill = fill
                                try:
                                    ws.cell(row=index+3,
                                            column=5,
                                            value=(a_result*100/no_employee)).fill = fill2
                                except:
                                    ws.cell(row=index+3,
                                            column=5,
                                            value=0).fill = fill2
                                try:
                                    ws.cell(row=index+4,
                                            column=5,
                                            value=(ma_result*100/no_employee)).fill = fill
                                except:
                                    ws.cell(row=index+4,
                                            column=5,
                                            value=0).fill = fill
                                if fill2 == blueFill:
                                    fill = blueFill
                                    fill2 = blue2Fill
                                else:
                                    fill = blue2Fill
                                    fill2 = blueFill
                    index += 5
                    int_emp += no_employee
            department_ids = False
        ws.cell(row=index, column=3, value='Total:').fill = blue3Fill
        ws.cell(row=index, column=4, value=int_emp).fill = blue3Fill
        return ws

    def open_print_report_demographic_xls(self):
        filename = 'Escuesta Demograficos.xlsx' 
        output = self._create_xlsx_demographic(
            Workbook, NamedTemporaryFile)
        xlsx = {
            'name': filename,
            # 'datas_fname': filename,
            'type': 'binary',
            'res_model': self._name,
            'datas': base64.b64encode(output),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        inserted_id = self.env['ir.attachment'].create(xlsx)
        url = '/web/content/%s?download=1' % (inserted_id.id)
        return {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': url
        }

    def _get_header_demographic(self):
        return [
            'DEMOGRAFICO',
            'SELECCIÓN',
            '#',
            '%',
            'Riesgo promedio',
        ]

    def _create_xlsx_demographic(self, Workbook, NamedTemporaryFile):
        wb = Workbook()
        ws = wb.active
        self._create_body_demographic(ws, self.user_input_ids)
        output = None
        with NamedTemporaryFile() as tmp:
            for column_cells in ws.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
            wb.save(tmp.name)
            output = tmp.read()
        return output

    def _create_body_demographic(self, ws, users):
        column = 1
        ws.cell(row=1,
                column=1,
                value='Demográficos - General').font = blueFont
        header = self._get_header_demographic()
        for c in header:
            cell = ws.cell(row=2, column=column, value=c)
            cell.fill = titleFill
            cell.font = whiteFont
            column += 1
        index = 3
        fill = blueFill
        fill2 = blue2Fill
        employees = self.env['hr.employee'].search([])
        departments = self.env['hr.department'].search([])
        job_ids = self.env['hr.job'].search([])
        int_job = len(job_ids)
        int_department = len(departments)

        # ws.merge_cells(start_row=index, start_column=1,
        #                 end_row=index+int_department-1, end_column=1)
        for row in range(index, index+int_department): 
            ws.cell(row=row,
                    column=1,
                    value='Departamento/ Sección/ Área').fill = fill
        index = index + int_department
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+11, end_column=1)
        for row in range(index, index+12):
            ws.cell(row=row,
                column=1,
                value='Edad en años').fill = fill2
        index = index + 12
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+4, end_column=1)
        for row in range(index, index+5):
            ws.cell(row=row,
                column=1,
                value='Estado Civil').fill = fill
        index = index + 5
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+7, end_column=1)
        for row in range(index, index+8):
            ws.cell(row=row,
                column=1,
                value='Nivel de estudios').fill = fill2
        index = index + 8
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+int_job-1, end_column=1)
        for row in range(index, index+int_job):
            ws.cell(row=row,
                column=1,
                value='Ocupación/ Profesión/ Puesto').fill = fill
        index = index + int_job
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+1, end_column=1)
        for row in range(index, index+2):
            ws.cell(row=row,
                column=1,
                value='Realiza rotación de turnos').fill = fill2
        index = index + 2
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+2, end_column=1)
        for row in range(index, index+3):
            ws.cell(row=row,
                column=1,
                value='Sexo').fill = fill
        index = index + 3
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+7, end_column=1)
        for row in range(index, index+8):
            ws.cell(row=row,
                column=1,
                value='Tiempo en el puesto actual').fill = fill2
        index = index + 8
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+7, end_column=1)
        for row in range(index, index+8):
            ws.cell(row=row,
                column=1,
                value='Tiempo de experiencia laboral').fill = fill
        index = index + 8
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+3, end_column=1)
        for row in range(index, index+4):
            ws.cell(row=row,
                column=1,
                value='Tipo de contratación').fill = fill2
        index = index + 4
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+2, end_column=1)
        for row in range(index, index+3):
            ws.cell(row=row,
                column=1,
                value='Tipo de jornada de trabajo').fill = fill
        index = index + 3
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+2, end_column=1)
        for row in range(index, index+3):
            ws.cell(row=row,
                column=1,
                value='Tipo de personal').fill = fill2
        index = index + 3
        # ws.merge_cells(start_row=index, start_column=1,
                        # end_row=index+3, end_column=1)
        for row in range(index, index+4):
            ws.cell(row=row,
                column=1,
                value='Tipo de puesto').fill = fill
        index = 3
        for department in departments:
            ws.cell(row=index, column=2, value=department.name).fill = fill
            employee_ids = self.env['hr.employee'].search([('department_id','=',department.id)])
            int_employee = len(employee_ids)
            ws.cell(row=index, column=3, value=int_employee).fill = fill
            cell = ws.cell(row=index, column=4, value=(int_employee*100/len(employees)))
            cell.fill = fill
            cell.number_format = "0.00"
            if int_employee > 0:
                usr = self.env['survey.user_input']
                for emp in employee_ids:
                    usr += self.env['survey.user_input'].search([('partner_id.name','=',emp.name)])
                result = 0.0
                if usr:
                    result += self._get_usr(usr)
                    ws.cell(row=index, column=5, value=int(result/len(usr))).fill = fill
                else:
                    ws.cell(row=index, column=5, value=0).fill = fill
            else:
                ws.cell(row=index, column=5, value=0).fill = fill
            index += 1
        ages = self._get_ages()
        for age in ages:
            ws.cell(row=index, column=2, value=age).fill = fill
            int_age = len(self.env['survey_score_template.survey_employee'].search([('year','=',age)]))
            ws.cell(row=index, column=3, value=int_age).fill = fill
            cell = ws.cell(row=index, column=4, value=(int_age*100/len(employees)))
            cell.fill = fill
            cell.number_format = "0.00"
            if int_employee > 0:
                employee_ids = self.env['survey_score_template.survey_employee'].search([('year','=',age)])
                usr = self.env['survey.user_input']
                for emp in employee_ids:
                    usr += self.env['survey.user_input'].search([('partner_id.name','=',emp.user_id.name)])
                result = 0.0
                if usr:
                    result += self._get_usr(usr)
                    ws.cell(row=index, column=5, value=int(result/len(usr))).fill = fill
                else:
                    ws.cell(row=index, column=5, value=0).fill = fill
            else:
                ws.cell(row=index, column=5, value=0).fill = fill
            index += 1
        sol_ids = self.env['hr.employee'].search([('marital','=','single')])
        sol = len(sol_ids)
        cas_ids = self.env['hr.employee'].search([('marital','=','married')])
        cas = len(cas_ids)
        coh_ids = self.env['hr.employee'].search([('marital','=','cohabitant')])
        coh = len(coh_ids)
        viu_ids = self.env['hr.employee'].search([('marital','=','widower')])
        viu = len(viu_ids)
        div_ids = self.env['hr.employee'].search([('marital','=','divorced')])
        div = len(div_ids)
        ws.cell(row=index, column=2, value="Soltero(a)").fill = fill
        value_sol = self._get_res_emp(sol_ids, sol)
        ws.cell(row=index+1, column=2, value="Casado(a)").fill = fill
        value_cas = self._get_res_emp(cas_ids, cas)
        ws.cell(row=index+2, column=2, value="Cohabitante").fill = fill
        value_coh = self._get_res_emp(coh_ids, coh)
        ws.cell(row=index+3, column=2, value="Viudo(a)").fill = fill
        value_viu = self._get_res_emp(viu_ids, viu)
        ws.cell(row=index+4, column=2, value="Divorciado(a)").fill = fill
        value_div = self._get_res_emp(div_ids, div)
        ws.cell(row=index, column=3, value=sol).fill = fill
        ws.cell(row=index+1, column=3, value=cas).fill = fill
        ws.cell(row=index+2, column=3, value=coh).fill = fill
        ws.cell(row=index+3, column=3, value=viu).fill = fill
        ws.cell(row=index+4, column=3, value=div).fill = fill
        cell = ws.cell(row=index, column=4, value=(sol*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(cas*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(coh*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+3, column=4, value=(viu*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+4, column=4, value=(div*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_sol).fill = fill
        ws.cell(row=index+1, column=5, value=value_cas).fill = fill
        ws.cell(row=index+2, column=5, value=value_coh).fill = fill
        ws.cell(row=index+3, column=5, value=value_viu).fill = fill
        ws.cell(row=index+4, column=5, value=value_div).fill = fill
        index += 5
        sf_ids = self.env['survey_score_template.survey_employee'].search([('level_sf','!=','')])
        sf = len(sf_ids)
        value_sf = self._get_res_usr(sf_ids, sf)
        primary_ids = self.env['survey_score_template.survey_employee'].search([('level_primary','=','Terminada')])
        primary = len(primary_ids)
        value_primary = self._get_res_usr(primary_ids, primary)
        secondary_ids = self.env['survey_score_template.survey_employee'].search([('level_secondary','=','Terminada')])
        secondary = len(secondary_ids)
        value_secondary = self._get_res_usr(secondary_ids, secondary)
        preparatory_ids = self.env['survey_score_template.survey_employee'].search([('level_preparatory','=','Terminada')])
        preparatory = len(preparatory_ids)
        value_preparatory = self._get_res_usr(preparatory_ids, preparatory)
        technical_ids = self.env['survey_score_template.survey_employee'].search([('level_technical','=','Terminada')])
        technical = len(technical_ids)
        value_technical = self._get_res_usr(technical_ids, technical)
        degree_ids = self.env['survey_score_template.survey_employee'].search([('level_degree','=','Terminada')])
        degree = len(degree_ids)
        value_degree = self._get_res_usr(degree_ids, degree)
        mastery_ids = self.env['survey_score_template.survey_employee'].search([('level_mastery','=','Terminada')])
        mastery = len(mastery_ids)
        value_mastery = self._get_res_usr(mastery_ids, mastery)
        doctorate_ids = self.env['survey_score_template.survey_employee'].search([('level_doctorate','=','Terminada')])
        doctorate = len(doctorate_ids)
        value_doctorate = self._get_res_usr(doctorate_ids, doctorate)
        ws.cell(row=index, column=2, value="Sin información").fill = fill
        ws.cell(row=index+1, column=2, value="Primaria").fill = fill
        ws.cell(row=index+2, column=2, value="Secundaria").fill = fill
        ws.cell(row=index+3, column=2, value="Preparatoria o Bachillerato").fill = fill
        ws.cell(row=index+4, column=2, value="Técnico Superior").fill = fill
        ws.cell(row=index+5, column=2, value="Licenciatura").fill = fill
        ws.cell(row=index+6, column=2, value="Maestria").fill = fill
        ws.cell(row=index+7, column=2, value="Doctorado").fill = fill
        ws.cell(row=index, column=3, value=sf).fill = fill
        ws.cell(row=index+1, column=3, value=primary).fill = fill
        ws.cell(row=index+2, column=3, value=secondary).fill = fill
        ws.cell(row=index+3, column=3, value=preparatory).fill = fill
        ws.cell(row=index+4, column=3, value=technical).fill = fill
        ws.cell(row=index+5, column=3, value=degree).fill = fill
        ws.cell(row=index+6, column=3, value=mastery).fill = fill
        ws.cell(row=index+7, column=3, value=doctorate).fill = fill
        cell = ws.cell(row=index, column=4, value=(sf*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(primary*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(secondary*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+3, column=4, value=(preparatory*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+4, column=4, value=(technical*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+5, column=4, value=(degree*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+6, column=4, value=(mastery*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+7, column=4, value=(doctorate*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_sf).fill = fill
        ws.cell(row=index+1, column=5, value=value_primary).fill = fill
        ws.cell(row=index+2, column=5, value=value_secondary).fill = fill
        ws.cell(row=index+3, column=5, value=value_preparatory).fill = fill
        ws.cell(row=index+4, column=5, value=value_technical).fill = fill
        ws.cell(row=index+5, column=5, value=value_degree).fill = fill
        ws.cell(row=index+6, column=5, value=value_mastery).fill = fill
        ws.cell(row=index+7, column=5, value=value_doctorate).fill = fill
        index += 8
        for job in job_ids:
            ws.cell(row=index, column=2, value=job.name).fill = fill
            jobs = self.env['hr.employee'].search([('job_id','=',job.id)])
            len_job = len(jobs)
            value_job = self._get_res_emp(jobs, len_job)
            ws.cell(row=index, column=3, value=len_job).fill = fill
            cell = ws.cell(row=index, column=4, value=(len_job*100/len(employees)))
            cell.fill = fill
            cell.number_format = "0.00"
            ws.cell(row=index, column=5, value=value_job).fill = fill
            index += 1
        turn_yes_ids = self.env['survey_score_template.survey_employee'].search([('turn','=','Sí')])
        turn_yes = len(turn_yes_ids)
        values_yes = self._get_res_usr(turn_yes_ids, turn_yes)
        turn_no_ids = self.env['survey_score_template.survey_employee'].search([('turn','=','No')])
        turn_no = len(turn_no_ids)
        values_no = self._get_res_usr(turn_no_ids, turn_no)
        ws.cell(row=index, column=2, value="Sí").fill = fill
        ws.cell(row=index+1, column=2, value="No").fill = fill
        ws.cell(row=index, column=3, value=turn_yes).fill = fill
        ws.cell(row=index+1, column=3, value=turn_no).fill = fill
        cell = ws.cell(row=index, column=4, value=(turn_yes*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(turn_no*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=values_yes).fill = fill
        ws.cell(row=index+1, column=5, value=values_no).fill = fill
        index += 2
        mas_ids = self.env['survey_score_template.survey_employee'].search([('sex','=','male')])
        mas = len(mas_ids)
        value_mas = self._get_res_usr(mas_ids, mas)
        fem_ids = self.env['survey_score_template.survey_employee'].search([('sex','=','female')])
        fem = len(fem_ids)
        value_fem = self._get_res_usr(fem_ids, fem)
        other_ids = self.env['survey_score_template.survey_employee'].search([('sex','=','other')])
        other = len(other_ids)
        value_other = self._get_res_usr(other_ids, other)
        ws.cell(row=index, column=2, value="Masculino").fill = fill
        ws.cell(row=index+1, column=2, value="Femenino").fill = fill
        ws.cell(row=index+2, column=2, value="Otro").fill = fill
        ws.cell(row=index, column=3, value=mas).fill = fill
        ws.cell(row=index+1, column=3, value=fem).fill = fill
        ws.cell(row=index+2, column=3, value=other).fill = fill
        cell = ws.cell(row=index, column=4, value=(mas*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(fem*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(other*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_mas).fill = fill
        ws.cell(row=index+1, column=5, value=value_fem).fill = fill
        ws.cell(row=index+2, column=5, value=value_other).fill = fill
        index += 3
        me6_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Menos de 6 meses')])
        ma1_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Entre 6 meses y  1 año')])
        a4_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Entre 1 a 4 años')])
        a9_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Entre 5 a 9 años')])
        a14_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Entre 10 a 14 años')])
        a19_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Entre 15 a 19 años')])
        a24_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','Entre 20 a 24 años')])
        ma25_ids = self.env['survey_score_template.survey_employee'].search([('experience_job','=','25 años o más')])
        me6 = len(me6_ids)
        ma1 = len(ma1_ids)
        a4 = len(a4_ids)
        a9 = len(a9_ids)
        a14 = len(a14_ids)
        a19 = len(a19_ids)
        a24 = len(a24_ids)
        ma25 = len(ma25_ids)
        value_me6 = self._get_res_usr(me6_ids, me6)
        value_a4 = self._get_res_usr(a4_ids, a4)
        value_a9 = self._get_res_usr(a9_ids, a9)
        value_a14 = self._get_res_usr(a14_ids, a14)
        value_ma1 = self._get_res_usr(ma1_ids, ma1)
        value_a19 = self._get_res_usr(a19_ids, a19)
        value_a24 = self._get_res_usr(a24_ids, a24)
        value_ma25 = self._get_res_usr(ma25_ids, ma25)
        ws.cell(row=index, column=2, value="Menos de 6 meses").fill = fill
        ws.cell(row=index+1, column=2, value="Entre 6 meses y 1 año").fill = fill
        ws.cell(row=index+2, column=2, value="Entre 1 a 4 años").fill = fill
        ws.cell(row=index+3, column=2, value="Entre 5 a 9 años").fill = fill
        ws.cell(row=index+4, column=2, value="Entre 10 a 14 años").fill = fill
        ws.cell(row=index+5, column=2, value="Entre 15 a 19 años").fill = fill
        ws.cell(row=index+6, column=2, value="Entre 20 a 24 años").fill = fill
        ws.cell(row=index+7, column=2, value="25 años o más").fill = fill
        ws.cell(row=index, column=3, value=me6).fill = fill
        ws.cell(row=index+1, column=3, value=ma1).fill = fill
        ws.cell(row=index+2, column=3, value=a4).fill = fill
        ws.cell(row=index+3, column=3, value=a9).fill = fill
        ws.cell(row=index+4, column=3, value=a14).fill = fill
        ws.cell(row=index+5, column=3, value=a19).fill = fill
        ws.cell(row=index+6, column=3, value=a24).fill = fill
        ws.cell(row=index+7, column=3, value=ma25).fill = fill
        cell = ws.cell(row=index, column=4, value=(me6*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(ma1*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(a4*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+3, column=4, value=(a9*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+4, column=4, value=(a14*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+5, column=4, value=(a19*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+6, column=4, value=(a24*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+7, column=4, value=(ma25*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=me6).fill = fill
        ws.cell(row=index+1, column=5, value=value_ma1).fill = fill
        ws.cell(row=index+2, column=5, value=value_a4).fill = fill
        ws.cell(row=index+3, column=5, value=value_a9).fill = fill
        ws.cell(row=index+4, column=5, value=value_a14).fill = fill
        ws.cell(row=index+5, column=5, value=value_a19).fill = fill
        ws.cell(row=index+6, column=5, value=value_a24).fill = fill
        ws.cell(row=index+7, column=5, value=value_ma25).fill = fill
        index += 8
        me6_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Menos de 6 meses')])
        ma1_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Entre 6 meses y  1 año')])
        a4_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Entre 1 a 4 años')])
        a9_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Entre 5 a 9 años')])
        a14_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Entre 10 a 14 años')])
        a19_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Entre 15 a 19 años')])
        a24_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','Entre 20 a 24 años')])
        ma25_work_ids = self.env['survey_score_template.survey_employee'].search([('experience_work','=','25 años o más')])
        me6_work = len(me6_work_ids)
        ma1_work = len(ma1_work_ids)
        a4_work = len(a4_work_ids)
        a9_work = len(a9_work_ids)
        a14_work = len(a14_work_ids)
        a19_work = len(a19_work_ids)
        a24_work = len(a24_work_ids)
        ma25_work = len(ma25_work_ids)
        value_me6_work = self._get_res_usr(me6_work_ids, me6_work)
        value_a4_work = self._get_res_usr(a4_work_ids, a4_work)
        value_a9_work = self._get_res_usr(a9_ids, a9_work)
        value_a14_work = self._get_res_usr(a14_work_ids, a14_work)
        value_ma1_work = self._get_res_usr(ma1_work_ids, ma1_work)
        value_a19_work = self._get_res_usr(a19_work_ids, a19_work)
        value_a24_work = self._get_res_usr(a24_work_ids, a24_work)
        value_ma25_work = self._get_res_usr(ma25_work_ids, ma25_work)
        ws.cell(row=index, column=2, value="Menos de 6 meses").fill = fill
        ws.cell(row=index+1, column=2, value="Entre 6 meses y 1 año").fill = fill
        ws.cell(row=index+2, column=2, value="Entre 1 a 4 años").fill = fill
        ws.cell(row=index+3, column=2, value="Entre 5 a 9 años").fill = fill
        ws.cell(row=index+4, column=2, value="Entre 10 a 14 años").fill = fill
        ws.cell(row=index+5, column=2, value="Entre 15 a 19 años").fill = fill
        ws.cell(row=index+6, column=2, value="Entre 20 a 24 años").fill = fill
        ws.cell(row=index+7, column=2, value="25 años o más").fill = fill
        ws.cell(row=index, column=3, value=me6_work).fill = fill
        ws.cell(row=index+1, column=3, value=ma1_work).fill = fill
        ws.cell(row=index+2, column=3, value=a4_work).fill = fill
        ws.cell(row=index+3, column=3, value=a9_work).fill = fill
        ws.cell(row=index+4, column=3, value=a14_work).fill = fill
        ws.cell(row=index+5, column=3, value=a19_work).fill = fill
        ws.cell(row=index+6, column=3, value=a24_work).fill = fill
        ws.cell(row=index+7, column=3, value=ma25_work).fill = fill
        cell = ws.cell(row=index, column=4, value=(me6_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(ma1_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(a4_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+3, column=4, value=(a9_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+4, column=4, value=(a14_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+5, column=4, value=(a19_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+6, column=4, value=(a24_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+7, column=4, value=(ma25_work*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=me6_work).fill = fill
        ws.cell(row=index+1, column=5, value=value_ma1_work).fill = fill
        ws.cell(row=index+2, column=5, value=value_a4_work).fill = fill
        ws.cell(row=index+3, column=5, value=a9_work).fill = fill
        ws.cell(row=index+4, column=5, value=value_a14_work).fill = fill
        ws.cell(row=index+5, column=5, value=value_a19_work).fill = fill
        ws.cell(row=index+6, column=5, value=value_a24_work).fill = fill
        ws.cell(row=index+7, column=5, value=value_ma25_work).fill = fill
        index += 8
        obra_ids = self.env['survey_score_template.survey_employee'].search([('contract_type','=','Por obra o proyecto')])
        obra = len(obra_ids)
        inde_ids = self.env['survey_score_template.survey_employee'].search([('contract_type','=','Tiempo indeterminado')])
        inde = len(inde_ids)
        dete_ids = self.env['survey_score_template.survey_employee'].search([('contract_type','=','Por tiempo determinado (temporal)')])
        dete = len(dete_ids)
        hono_ids = self.env['survey_score_template.survey_employee'].search([('contract_type','=','Honorarios')])
        hono = len(hono_ids)
        value_obra = self._get_res_usr(obra_ids, obra)
        value_inde = self._get_res_usr(inde_ids, inde)
        value_dete = self._get_res_usr(dete_ids, dete)
        value_hono = self._get_res_usr(hono_ids, hono)
        ws.cell(row=index, column=2, value="Por obra o proyecto").fill = fill
        ws.cell(row=index+1, column=2, value="Tiempo indeterminado").fill = fill
        ws.cell(row=index+2, column=2, value="Por tiempo determinado (temporal)").fill = fill
        ws.cell(row=index+3, column=2, value="Honorarios").fill = fill
        ws.cell(row=index, column=3, value=obra).fill = fill
        ws.cell(row=index+1, column=3, value=inde).fill = fill
        ws.cell(row=index+2, column=3, value=dete).fill = fill
        ws.cell(row=index+3, column=3, value=hono).fill = fill
        cell = ws.cell(row=index, column=4, value=(obra*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(inde*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(dete*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+3, column=4, value=(hono*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_obra).fill = fill
        ws.cell(row=index+1, column=5, value=value_inde).fill = fill
        ws.cell(row=index+2, column=5, value=value_dete).fill = fill
        ws.cell(row=index+3, column=5, value=value_hono).fill = fill
        index += 4
        noc_ids = self.env['survey_score_template.survey_employee'].search([('jornal_type','=','Fijo nocturno (entre las 20:00 y 6:00 hrs)')])
        mix_ids = self.env['survey_score_template.survey_employee'].search([('jornal_type','=','Fijo mixto (combinación de nocturno y diurno)')])
        diu_ids = self.env['survey_score_template.survey_employee'].search([('jornal_type','=','Fijo diurno (entre las 6:00 y 20:00 hrs)')])
        noc = len(noc_ids)
        mix = len(mix_ids)
        diu = len(diu_ids)
        value_noc = self._get_res_usr(noc_ids, noc)
        value_mix = self._get_res_usr(mix_ids, mix)
        value_diu = self._get_res_usr(diu_ids, diu)
        ws.cell(row=index, column=2, value="Fijo nocturno (entre las 20:00 y 6:00 hrs)").fill = fill
        ws.cell(row=index+1, column=2, value="Fijo mixto (combinación de nocturno y diurno)").fill = fill
        ws.cell(row=index+2, column=2, value="Fijo diurno (entre las 6:00 y 20:00 hrs)").fill = fill
        ws.cell(row=index, column=3, value=noc).fill = fill
        ws.cell(row=index+1, column=3, value=mix).fill = fill
        ws.cell(row=index+2, column=3, value=diu).fill = fill
        cell = ws.cell(row=index, column=4, value=(noc*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(mix*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(diu*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_noc).fill = fill
        ws.cell(row=index+1, column=5, value=value_mix).fill = fill
        ws.cell(row=index+2, column=5, value=value_diu).fill = fill
        index += 3
        sindi_ids = self.env['survey_score_template.survey_employee'].search([('personal_type','=','Sindicalizado')])
        sindi = len(sindi_ids)
        value_sindi = self._get_res_usr(sindi_ids, sindi)
        confi_ids = self.env['survey_score_template.survey_employee'].search([('personal_type','=','Confianza')])
        confi = len(confi_ids)
        value_confi = self._get_res_usr(confi_ids, confi)
        nunguno_ids = self.env['survey_score_template.survey_employee'].search([('personal_type','=','Ninguno')])
        nunguno = len(nunguno_ids)
        value_nunguno = self._get_res_usr(nunguno_ids, nunguno)
        ws.cell(row=index, column=2, value="Sindicalizado").fill = fill
        ws.cell(row=index+1, column=2, value="Confianza").fill = fill
        ws.cell(row=index+2, column=2, value="Ninguno").fill = fill
        ws.cell(row=index, column=3, value=sindi).fill = fill
        ws.cell(row=index+1, column=3, value=confi).fill = fill
        ws.cell(row=index+2, column=3, value=nunguno).fill = fill
        cell = ws.cell(row=index, column=4, value=(sindi*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(confi*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(nunguno*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_sindi).fill = fill
        ws.cell(row=index+1, column=5, value=value_confi).fill = fill
        ws.cell(row=index+2, column=5, value=value_nunguno).fill = fill
        index += 3
        opera_ids = self.env['survey_score_template.survey_employee'].search([('job_type','=','Operativo')])
        opera = len(opera_ids)
        sup_ids = self.env['survey_score_template.survey_employee'].search([('job_type','=','Supervisor')])
        sup = len(sup_ids)
        pro_ids = self.env['survey_score_template.survey_employee'].search([('job_type','=','Profesional o técnico')])
        pro = len(pro_ids)
        gerente_ids = self.env['survey_score_template.survey_employee'].search([('job_type','=','Gerente')])
        gerente = len(gerente_ids)
        value_opera = self._get_res_usr(opera_ids, opera)
        value_sup = self._get_res_usr(sup_ids, sup)
        value_pro = self._get_res_usr(pro_ids, pro)
        value_gerente = self._get_res_usr(gerente_ids, gerente)
        ws.cell(row=index, column=2, value="Operativo").fill = fill
        ws.cell(row=index+1, column=2, value="Supervisor").fill = fill
        ws.cell(row=index+2, column=2, value="Profesional o técnico").fill = fill
        ws.cell(row=index+3, column=2, value="Gerente").fill = fill
        ws.cell(row=index, column=3, value=opera).fill = fill
        ws.cell(row=index+1, column=3, value=sup).fill = fill
        ws.cell(row=index+2, column=3, value=pro).fill = fill
        ws.cell(row=index+3, column=3, value=gerente).fill = fill
        cell = ws.cell(row=index, column=4, value=(opera*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+1, column=4, value=(sup*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+2, column=4, value=(pro*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        cell = ws.cell(row=index+3, column=4, value=(gerente*100/len(employees)))
        cell.fill = fill
        cell.number_format = "0.00"
        ws.cell(row=index, column=5, value=value_opera).fill = fill
        ws.cell(row=index+1, column=5, value=value_sup).fill = fill
        ws.cell(row=index+2, column=5, value=value_pro).fill = fill
        ws.cell(row=index+3, column=5, value=value_gerente).fill = fill
        index += 4
        return ws

    def _get_ages(self):
        return [
            '15-19',
            '20-24',
            '25-29',
            '30-34',
            '35-39',
            '40-44',
            '45-49',
            '50-54',
            '55-59',
            '60-64',
            '65-69',
            '70 o mas',
        ]


    def _get_fill(self, fill):
        if fill == blueFill:
            fill = blue2Fill
        else:
            fill = blueFill
        return fill

class SurveyInvite(models.TransientModel):
    _inherit = 'survey.invite'

    survey_url = fields.Char(related="survey_id.survey_url", readonly=True)