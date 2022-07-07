# _*_ coding:utf_8 _*_
from odoo import models, fields, api
from datetime import date, datetime
import base64
from base64 import b64encode
from werkzeug import urls

from tabulate import tabulate
import openpyxl
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from pandas import DataFrame, crosstab
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment

titleFill = PatternFill(start_color='000CA4', end_color='000CA4', fill_type='solid')
blueFill = PatternFill(start_color='8AEAFB', end_color='8AEAFB', fill_type='solid')
blue2Fill = PatternFill(start_color='3DE2FF', end_color='3DE2FF', fill_type='solid')
blue3Fill = PatternFill(start_color='8087FF', end_color='8087FF', fill_type='solid')
blue4Fill = PatternFill(start_color='33A8FF', end_color='33A8FF', fill_type='solid')
lowFill = PatternFill(start_color='DAF7A6', end_color='DAF7A6', fill_type='solid')
mediumFill = PatternFill(start_color='FFC300', end_color='FFC300', fill_type='solid')
highFill = PatternFill(start_color='FF5733', end_color='FF5733', fill_type='solid')
veryFill = PatternFill(start_color='C70039', end_color='C70039', fill_type='solid')
blueFont = Font(name='Arial', size=15, color='000CA4')
whiteFont = Font(color='FFFFFF')
alig = Alignment(horizontal="center")
ver_al = al = Alignment(vertical="center")

class SurveyWorkplace(models.Model):
    _name = 'survey.workplace'

    name = fields.Char(string='Centro de trabajo', required=True)

class SurveySurvey(models.Model):
    _inherit = 'survey.survey'

    def open_print_report_workplace_xls(self):
        filename = 'Centros de trabajo.xlsx' 
        output = self._create_xlsx_workplace(
            Workbook, NamedTemporaryFile)
        xlsx = {
            'name': filename,
            # 'datas_fname': filename,
            'type': 'binary',
            'res_model': self._name,
            'datas': base64.b64encode(output),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        inserted_id = self.env['ir.attachment'].create(xlsx)
        url = '/web/content/%s?download=1' % (inserted_id.id)
        return {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': url
        }

    def _get_header_workplace(self):
        return [
            'Centro de trabajo',
            'Departamento',
            'Categoría',
            'Valor',
            'Dominio',
            'Valor',
        ]

    def _create_xlsx_workplace(self, Workbook, NamedTemporaryFile):
        wb = Workbook()
        ws = wb.active
        self._create_body_workplace(ws, self.user_input_ids)
        output = None
        with NamedTemporaryFile() as tmp:
            for column_cells in ws.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
            wb.save(tmp.name)
            output = tmp.read()
        return output

    def _create_body_workplace(self, ws, users):
        column = 1
        ws.cell(row=1,
                column=1,
                value='Resultados Categorías / Dominios').font = blueFont
        header = self._get_header_workplace()
        for c in header:
            cell = ws.cell(row=2, column=column, value=c)
            cell.fill = titleFill
            cell.font = whiteFont
            column += 1
        index = 3
        index3 = 3
        fill = blueFill
        fill2 = blue2Fill
        fill3 = blueFill
        fill4 = blueFill
        workplace_ids = self.env['survey.workplace']
        employee_ids = self.env['hr.employee']
        categ = self._get_categ_data()
        dom = self._get_dom()
        for user in users:
            employee_ids += self.env['hr.employee'].search([('name','=',user.partner_id.name)])
            for emp in employee_ids:
                if not emp.workplace_id in workplace_ids:
                    workplace_ids += emp.workplace_id
        int_emp = 0
        for workplace in workplace_ids:
            department_ids = self.env['hr.department']
            for emp in employee_ids:
                if not emp.department_id in department_ids and emp.workplace_id.id == workplace.id:
                    department_ids += emp.department_id
            if department_ids:
                # ws.merge_cells(start_row=index, start_column=1,
                #                 end_row=index+(len(department_ids)*10)-1, end_column=1)
                for row in range(index, index+(len(department_ids)*10)):
                    cell = ws.cell(row=row,
                                    column=1,
                                    value=workplace.name)
                    cell.fill = fill
                    cell.alignment = ver_al
                for department in department_ids:
                    # ws.merge_cells(start_row=index3, start_column=2,
                    #                 end_row=index3+9, end_column=2)
                    for row in range(index3, index3+10):
                        cell = ws.cell(row=row,
                                    column=2,
                                    value=department.name)
                        cell.fill = fill
                        cell.alignment = ver_al
                    index2 = 0
                    for c in categ:
                        val = 0.00
                        ws.merge_cells(start_row=index3+index2, start_column=3,
                                    end_row=index3+index2+1, end_column=3)
                        ws.cell(row=index3+index2, column=3, value=c).fill = fill3
                        for emp in employee_ids:
                            for user in users:
                                if emp.department_id.id == department.id and emp.name == user.partner_id.name:
                                    for line in user.user_input_line_ids:
                                        if line.value_suggested_row.category_id.name == c:
                                            val += line.answer_score
                        ws.merge_cells(start_row=index3+index2, start_column=4,
                                    end_row=index3+index2+1, end_column=4)
                        ws.cell(row=index3+index2, column=4, value=val).fill = fill3
                        fill3 = self._get_fill(fill3)
                        index2 += 2
                    index2 = 0
                    for d in dom:
                        val = 0.00
                        ws.cell(row=index3+index2, column=5, value=d).fill = fill4
                        for emp in employee_ids:
                            for user in users:
                                if emp.department_id.id == department.id and emp.name == user.partner_id.name:
                                    for line in user.user_input_line_ids:
                                        if line.value_suggested_row.domain_id.name == d:
                                            val += line.answer_score
                        ws.cell(row=index3+index2, column=6, value=val).fill = fill4
                        fill4 = self._get_fill(fill4)
                        index2 += 1
                    index3 += 10
                index = index+(len(department_ids)*10)
                fill = self._get_fill(fill)
        return ws