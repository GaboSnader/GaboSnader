# _*_ coding:utf_8 _*_
from odoo import models, fields, api
from datetime import date, datetime
import base64
from base64 import b64encode
from werkzeug import urls

from tabulate import tabulate
import openpyxl
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from pandas import DataFrame, crosstab
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment

titleFill = PatternFill(start_color='000CA4', end_color='000CA4', fill_type='solid')
blueFill = PatternFill(start_color='8AEAFB', end_color='8AEAFB', fill_type='solid')
blue2Fill = PatternFill(start_color='3DE2FF', end_color='3DE2FF', fill_type='solid')
blue3Fill = PatternFill(start_color='8087FF', end_color='8087FF', fill_type='solid')
blue4Fill = PatternFill(start_color='33A8FF', end_color='33A8FF', fill_type='solid')
lowFill = PatternFill(start_color='DAF7A6', end_color='DAF7A6', fill_type='solid')
mediumFill = PatternFill(start_color='FFC300', end_color='FFC300', fill_type='solid')
highFill = PatternFill(start_color='FF5733', end_color='FF5733', fill_type='solid')
veryFill = PatternFill(start_color='C70039', end_color='C70039', fill_type='solid')
blueFont = Font(name='Arial', size=15, color='000CA4')
whiteFont = Font(color='FFFFFF')
alig = Alignment(horizontal="center")
ver_al = al = Alignment(vertical="center")

class SurveySurvey(models.Model):
    _inherit = 'survey.survey'

    def open_print_report_traumatic_xls(self):
        filename = 'Acontecimientos traumáticos.xlsx' 
        output = self._create_xlsx_traumatic(
            Workbook, NamedTemporaryFile)
        xlsx = {
            'name': filename,
            # 'datas_fname': filename,
            'type': 'binary',
            'res_model': self._name,
            'datas': base64.b64encode(output),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        inserted_id = self.env['ir.attachment'].create(xlsx)
        url = '/web/content/%s?download=1' % (inserted_id.id)
        return {
            'type': 'ir.actions.act_url',
            'name': filename,
            'url': url
        }

    def _get_header_traumatic(self):
        return [
            'CENTRO DE TRABAJO',
            'DEPARTAMENTO',
            'NOMBRE COLABORADOR',
            'PREGUNTA',
            'DETALLE COLABORADOR',
            'FUERON SUJETOS A ACONTECIMIENTOS TRAUMATICOS',
            'REQUIERE ATENCIÓN CLINICA',
        ]

    def _create_xlsx_traumatic(self, Workbook, NamedTemporaryFile):
        wb = Workbook()
        ws = wb.active
        self._create_body_traumatic(ws, self.user_input_ids)
        output = None
        with NamedTemporaryFile() as tmp:
            for column_cells in ws.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
            wb.save(tmp.name)
            output = tmp.read()
        return output

    def _get_traumatic(self, trauma):
        tr = 0
        if trauma.question_7 or trauma.question_8:
            return True
        else:
            if trauma.question_9:
                tr += 1
            if trauma.question_10:
                tr += 1
            if trauma.question_11:
                tr += 1
            if trauma.question_12:
                tr += 1
            if trauma.question_13:
                tr += 1
            if trauma.question_14:
                tr += 1
            if trauma.question_15:
                tr += 1
            if trauma.question_16:
                tr += 1
            if tr >= 3:
                return True
            else:
                tr = 0
                if trauma.question_17:
                    tr += 1
                if trauma.question_18:
                    tr += 1
                if trauma.question_19:
                    tr += 1
                if trauma.question_20:
                    tr += 1
                if tr >= 2:
                    return True
                else:
                    return False


    def _create_body_traumatic(self, ws, users):
        column = 1
        ws.cell(row=1,
                column=1,
                value='Acontecimientos - General').font = blueFont
        header = self._get_header_traumatic()
        for c in header:
            cell = ws.cell(row=2, column=column, value=c)
            cell.fill = titleFill
            cell.font = whiteFont
            column += 1
        index = 3
        index2 = 3
        fill = blueFill
        fill2 = blue2Fill
        fill3 = blueFill
        survey_employee_id = self.env['survey_score_template.survey_employee'].search([])
        for emp in survey_employee_id:
            for trauma in emp.traumatic:
                if trauma.question_6 or trauma.question_5 or trauma.question_4 or trauma.question_3 or trauma.question_2 or trauma.question_1:
                    ws.cell(row=index, column=6, value='Si').fill = blue3Fill
                    traumatic_bol = self._get_traumatic(trauma)
                    if traumatic_bol:
                        cell = ws.cell(row=index, column=7, value='Si')
                        cell.alignment = alig
                        cell.fill = blue3Fill
                    else:
                        cell = ws.cell(row=index, column=7, value='No')
                        cell.alignment = alig
                        cell.fill = blue3Fill
                else:
                    cell = ws.cell(row=index, column=6, value='No')
                    cell.alignment = alig
                    cell.fill = blue3Fill
                    cell = ws.cell(row=index, column=7, value='No')
                    cell.alignment = alig
                    cell.fill = blue3Fill
                ws.cell(row=index, column=4, value='Acontecimiento traumatico').fill = blue4Fill
                ws.cell(row=index, column=5).fill = blue4Fill
                ws.cell(row=index, column=6).fill = blue4Fill
                ws.cell(row=index, column=7).fill = blue4Fill
                ws.cell(row=index+1, column=4, 
                    value='¿Accidente que tenga como consecuencia la muerte, la pérdida de un miembro o una lesión grave?').fill = fill
                ws.cell(row=index+1, column=5, value='Si' if trauma.question_1 else 'No').fill = fill
                ws.cell(row=index+1, column=6).fill = fill
                ws.cell(row=index+1, column=7).fill = fill
                ws.cell(row=index+2, column=4, 
                    value='¿Asaltos?').fill = fill2
                ws.cell(row=index+2, column=5, value='Si' if trauma.question_2 else 'No').fill = fill2
                ws.cell(row=index+2, column=6).fill = fill2
                ws.cell(row=index+2, column=7).fill = fill2
                ws.cell(row=index+3, column=4, 
                    value='¿Actos violentos que derivaron en lesiones graves?').fill = fill
                ws.cell(row=index+3, column=5, value='Si' if trauma.question_3 else 'No').fill = fill
                ws.cell(row=index+3, column=6).fill = fill
                ws.cell(row=index+3, column=7).fill = fill
                ws.cell(row=index+4, column=4, 
                    value='¿Secuestro?').fill = fill2
                ws.cell(row=index+4, column=5, value='Si' if trauma.question_4 else 'No').fill = fill2
                ws.cell(row=index+4, column=6).fill = fill2
                ws.cell(row=index+4, column=7).fill = fill2
                ws.cell(row=index+5, column=4, 
                    value='¿Amenazas?').fill = fill
                ws.cell(row=index+5, column=5, value='Si' if trauma.question_5 else 'No').fill = fill
                ws.cell(row=index+5, column=6).fill = fill
                ws.cell(row=index+5, column=7).fill = fill
                ws.cell(row=index+6, column=4, 
                    value='¿Cualquier otro que ponga en riesgo su vida o salud, y/o la de otras personas?').fill = fill2
                ws.cell(row=index+6, column=5, value='Si' if trauma.question_6 else 'No').fill = fill2
                ws.cell(row=index+6, column=6).fill = fill2
                ws.cell(row=index+6, column=7).fill = fill2
                index += 7
                if trauma.question_7 or trauma.question_8:
                    ws.cell(row=index, column=4, value='Recuerdos persistentes sobre el acontecimiento').fill = blue4Fill
                    ws.cell(row=index, column=5).fill = blue4Fill
                    ws.cell(row=index, column=6).fill = blue4Fill
                    ws.cell(row=index, column=7).fill = blue4Fill
                    ws.cell(row=index+1, column=4, 
                        value='¿Ha tenido recuerdos recurrentes sobre el acontecimiento que le provocan malestares?').fill = fill
                    ws.cell(row=index+1, column=5, value='Si' if trauma.question_7 else 'No').fill = fill
                    ws.cell(row=index+1, column=6).fill = fill
                    ws.cell(row=index+1, column=7).fill = fill
                    ws.cell(row=index+2, column=4, 
                        value='¿Ha tenido sueños de carácter recurrente sobre el acontecimiento, que le producen malestar?').fill = fill2
                    ws.cell(row=index+2, column=5, value='Si' if trauma.question_8 else 'No').fill = fill2
                    ws.cell(row=index+2, column=6).fill = fill2
                    ws.cell(row=index+2, column=7).fill = fill2
                    index += 3
                if trauma.question_9 or trauma.question_10 or trauma.question_11 or trauma.question_12 or trauma.question_13 or trauma.question_14 or trauma.question_15 or trauma.question_16:
                    ws.cell(row=index, column=4, value='Esfuerzo por evitar circunstancias parecidas o asociadas al acontecimiento').fill = blue4Fill
                    ws.cell(row=index, column=5).fill = blue4Fill
                    ws.cell(row=index, column=6).fill = blue4Fill
                    ws.cell(row=index, column=7).fill = blue4Fill
                    ws.cell(row=index+1, column=4, 
                    value='¿Se ha esforzado por evitar todo tipo de sentimientos, conversaciones o situaciones que le puedan recordar el acontecimiento?').fill = fill
                    ws.cell(row=index+1, column=5, value='Si' if trauma.question_9 else 'No').fill = fill
                    ws.cell(row=index+1, column=6).fill = fill
                    ws.cell(row=index+1, column=7).fill = fill
                    ws.cell(row=index+2, column=4, 
                    value='¿Se ha esforzado por evitar todo tipo de actividades, lugares o personas que motivan recuerdos del acontecimiento?').fill = fill2
                    ws.cell(row=index+2, column=5, value='Si' if trauma.question_10 else 'No').fill = fill2
                    ws.cell(row=index+2, column=6).fill = fill2
                    ws.cell(row=index+2, column=7).fill = fill2
                    ws.cell(row=index+3, column=4, 
                    value='¿Ha tenido dificultad para recordar alguna parte importante del evento?').fill = fill
                    ws.cell(row=index+3, column=5, value='Si' if trauma.question_11 else 'No').fill = fill
                    ws.cell(row=index+3, column=6).fill = fill
                    ws.cell(row=index+3, column=7).fill = fill
                    ws.cell(row=index+4, column=4, 
                    value='¿Ha disminuido su interés en sus actividades cotidianas?').fill = fill2
                    ws.cell(row=index+4, column=5, value='Si' if trauma.question_12 else 'No').fill = fill2
                    ws.cell(row=index+4, column=6).fill = fill2
                    ws.cell(row=index+4, column=7).fill = fill2
                    ws.cell(row=index+5, column=4, 
                    value='¿Se ha sentido usted alejado o distante de los demás?').fill = fill
                    ws.cell(row=index+5, column=5, value='Si' if trauma.question_13 else 'No').fill = fill
                    ws.cell(row=index+5, column=6).fill = fill
                    ws.cell(row=index+5, column=7).fill = fill
                    ws.cell(row=index+6, column=4, 
                    value='¿Ha notado que tiene dificultad para expresar sus sentimientos?').fill = fill2
                    ws.cell(row=index+6, column=5, value='Si' if trauma.question_14 else 'No').fill = fill2
                    ws.cell(row=index+6, column=6).fill = fill2
                    ws.cell(row=index+6, column=7).fill = fill2
                    ws.cell(row=index+7, column=4, 
                    value='¿Ha tenido la impresión de que su vida se va a acortar, que va a morir antes que otras personas o que tiene un futuro limitado?').fill = fill
                    ws.cell(row=index+7, column=5, value='Si' if trauma.question_15 else 'No').fill = fill
                    ws.cell(row=index+7, column=6).fill = fill
                    ws.cell(row=index+7, column=7).fill = fill
                    ws.cell(row=index+8, column=4, 
                    value='¿Ha tenido usted dificultades para dormir?').fill = fill2
                    ws.cell(row=index+8, column=5, value='Si' if trauma.question_16 else 'No').fill = fill2
                    ws.cell(row=index+8, column=6).fill = fill2
                    ws.cell(row=index+8, column=7).fill = fill2
                    index += 9
                if trauma.question_17 or trauma.question_18 or trauma.question_19 or trauma.question_20:
                    ws.cell(row=index, column=4, value='Afectación').fill = blue4Fill
                    ws.cell(row=index, column=5).fill = blue4Fill
                    ws.cell(row=index, column=6).fill = blue4Fill
                    ws.cell(row=index, column=7).fill = blue4Fill
                    ws.cell(row=index+1, column=4, 
                        value='¿Ha estado particularmente irritable o le han dado arranques de coraje?').fill = fill
                    ws.cell(row=index+1, column=5, value='Si' if trauma.question_17 else 'No').fill = fill
                    ws.cell(row=index+1, column=6).fill = fill
                    ws.cell(row=index+1, column=7).fill = fill
                    ws.cell(row=index+2, column=4, 
                        value='¿Ha tenido dificultad para concentrarse?').fill = fill2
                    ws.cell(row=index+2, column=5, value='Si' if trauma.question_18 else 'No').fill = fill2
                    ws.cell(row=index+2, column=6).fill = fill2
                    ws.cell(row=index+2, column=7).fill = fill2
                    ws.cell(row=index+3, column=4, 
                        value='¿Ha estado nervioso o constantemente en alerta?').fill = fill
                    ws.cell(row=index+3, column=5, value='Si' if trauma.question_19 else 'No').fill = fill
                    ws.cell(row=index+3, column=6).fill = fill
                    ws.cell(row=index+3, column=7).fill = fill
                    ws.cell(row=index+4, column=4, 
                        value='¿Se ha sobresaltado fácilmente por cualquier cosa?').fill = fill2
                    ws.cell(row=index+4, column=5, value='Si' if trauma.question_20 else 'No').fill = fill2
                    ws.cell(row=index+4, column=6).fill = fill2
                    ws.cell(row=index+4, column=7).fill = fill2
                    index += 5
                department = ''
                workplace = ''
                employee_ids = self.env['hr.employee'].search([('name','=',emp.user_id.name)])
                for emp2 in employee_ids:
                    department = str(emp2.department_id.name)
                    workplace = str(emp2.workplace_id.name)
                for row in range(index2, index):
                    cell = ws.cell(row=row, column=1, value=department)
                    cell.alignment = ver_al
                    cell.fill = fill3
                # ws.merge_cells(start_row=index2, start_column=1,
                #                 end_row=index-1, end_column=1)
                for row in range(index2, index):
                    cell = ws.cell(row=row, column=2, value=workplace)
                    cell.alignment = ver_al
                    cell.fill = fill3
                # ws.merge_cells(start_row=index2, start_column=2,
                #                 end_row=index-1, end_column=2)
                for row in range(index2, index):
                    cell = ws.cell(row=row, column=3, value=emp.user_id.name)
                    cell.alignment = ver_al
                    cell.fill = fill3
                # ws.merge_cells(start_row=index2, start_column=3,
                #                 end_row=index-1, end_column=3)
                fill3 = self._get_fill(fill3)
                index2 = index
        ws.cell(row=index,
                column=1,
                value='Total').fill = blue3Fill
        ws.cell(row=index,
                column=2,
                value=self.answer_done_count).fill = blue3Fill
        return ws

    def _get_fill(self, fill):
        if fill == blueFill:
            fill = blue2Fill
        else:
            fill = blueFill
        return fill
